"""excel_report 超链接回归测试（W8，对齐 misteryanalyze fa1444ff）。

验证：
- 汇总报告「代码」列超链接 → 对应个股 sheet A1（location 内部引用）
- 个股 sheet 第 1 行导航：首页(→汇总报告) / 前一页 / 后一页
- 所有超链接 location 目标 sheet 真实存在（无死链）
"""
from __future__ import annotations

import io

import openpyxl
import pytest

from mystery.apps.reports.excel_report import excel_bytes


def _mk(symbol: str, name: str, score: float) -> dict:
    return {
        "symbol": symbol, "name": name, "trade_date": "2026-08-30",
        "score": score, "advice": "持有", "true_resonance": True, "price": 10.0,
        "mystery": {"vap_atr": {"POC": 9.0}, "platform": {"平台状态": "平台"},
                    "main_wave": {"主升浪状态": "主升浪"}, "resonance": {},
                    "checklist8": {}, "signal": {}, "technical": {}},
        "financial": {}, "sector": {"行业名称": "测试"}, "chan": {},
    }


@pytest.fixture()
def wb():
    """乱序输入（验证排序不影响超链接匹配）。"""
    rs = [_mk("sh600519", "贵州茅台", 49.0),
          _mk("sz000001", "平安银行", 70.0),
          _mk("sh600150", "中国船舶", 20.0)]
    data = excel_bytes(rs)
    return openpyxl.load_workbook(io.BytesIO(data))


def test_summary_code_hyperlink_to_detail(wb):
    """汇总报告代码列超链接 location 指向对应个股 sheet。"""
    ws = wb["汇总报告"]
    expect = {
        "sz000001": "个股平安银行_sz000001",
        "sh600519": "个股贵州茅台_sh600519",
        "sh600150": "个股中国船舶_sh600150",
    }
    found = {}
    for row in ws.iter_rows(min_row=2):
        c = row[0]
        assert c.hyperlink is not None, f"{c.value} 代码列缺超链接"
        # location 内部引用（target 为 None，不会补全文件路径）
        assert c.hyperlink.target is None
        assert c.hyperlink.location == f"'{expect[c.value]}'!A1"
        found[c.value] = True
    assert set(found) == set(expect)


def test_detail_navigation_first_last(wb):
    """首/末个股 sheet 只有首页 + 后一页/前一页。"""
    first = wb["个股贵州茅台_sh600519"]
    last = wb["个股中国船舶_sh600150"]
    # 第一个：首页 + 后一页，无前一页
    assert first.cell(row=1, column=1).value == "首页"
    assert first.cell(row=1, column=1).hyperlink.location == "'汇总报告'!A1"
    assert first.cell(row=1, column=2).value is None          # 无前一页
    assert first.cell(row=1, column=3).value == "后一页"
    assert (first.cell(row=1, column=3).hyperlink.location
            == "'个股平安银行_sz000001'!A1")
    # 最后一个：首页 + 前一页，无后一页
    assert last.cell(row=1, column=1).value == "首页"
    assert last.cell(row=1, column=2).value == "前一页"
    assert (last.cell(row=1, column=2).hyperlink.location
            == "'个股平安银行_sz000001'!A1")
    assert last.cell(row=1, column=3).value is None           # 无后一页


def test_detail_navigation_middle(wb):
    """中间个股 sheet 含 首页/前一页/后一页。"""
    mid = wb["个股平安银行_sz000001"]
    assert mid.cell(row=1, column=1).value == "首页"
    assert (mid.cell(row=1, column=1).hyperlink.location == "'汇总报告'!A1")
    assert mid.cell(row=1, column=2).value == "前一页"
    assert (mid.cell(row=1, column=2).hyperlink.location
            == "'个股贵州茅台_sh600519'!A1")
    assert mid.cell(row=1, column=3).value == "后一页"
    assert (mid.cell(row=1, column=3).hyperlink.location
            == "'个股中国船舶_sh600150'!A1")


def test_no_dead_links(wb):
    """所有超链接 location 目标 sheet 必须存在。"""
    for sname in wb.sheetnames:
        w = wb[sname]
        for row in w.iter_rows():
            for c in row:
                if c.hyperlink and c.hyperlink.location:
                    target = c.hyperlink.location.split("'")[1]
                    assert target in wb.sheetnames, f"{sname}!{c.coordinate} 死链 -> {target}"


def test_detail_data_starts_row2(wb):
    """个股明细数据从第 2 行开始（第 1 行预留给导航）。"""
    w = wb["个股贵州茅台_sh600519"]
    assert w.cell(row=2, column=1).value == "项目"
    assert w.cell(row=2, column=2).value == "结果"
    assert w.cell(row=2, column=3).value == "备注"
